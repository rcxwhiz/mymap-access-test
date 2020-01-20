from __future__ import annotations

import pickle
import os
import sys
import openpyxl
from pathlib import Path
from typing import Optional
from classroom import semester_getter
import classroom


class SemesterManagerMeta(type):
	_instance: Optional[SemesterManager] = None

	def __call__(self) -> SemesterManager:
		if self._instance is None:
			self._instance = super().__call__()
		return self._instance


class SemesterManager(metaclass=SemesterManagerMeta):

	def __init__(self):
		self.semesters = {}

		self.pickles_path = os.path.join(os.path.dirname(Path(sys.argv[0]).parent), 'pickles')
		if not os.path.exists(self.pickles_path):
			os.makedirs(self.pickles_path)

		for file in os.listdir(self.pickles_path):
			self.semesters[file] = pickle.load(open(os.path.join(self.pickles_path, file), 'rb'))

		self.selected_semester = None
		self.filtered_sections = []

		self.dept_filter = ''
		self.course_num_filter = ''
		self.course_name_filter = ''
		self.instructor_filter = ''
		self.time_filter = [0, 0]
		self.type_filter = {'DAY': False,
		                    'EVENING': False,
		                    'ONLINE': False,
		                    'SALT LAKE': False,
		                    'ST ABROAD': False}
		self.day_filter = {'M': [0, 0, False],
		                   'T': [0, 0, False],
		                   'W': [0, 0, False],
		                   'Th': [0, 0, False],
		                   'F': [0, 0, False],
		                   'Sa': [0, 0, False]}
		self.credits_filter = [0.0, 0.0]  # MIN - MAX
		self.course_level_filter = {100: False,
		                            200: False,
		                            300: False,
		                            400: False,
		                            500: False,
		                            600: False,
		                            700: False}

	def save_to_xlsx(self):
		wb = openpyxl.Workbook()

		for i, semester_year in enumerate(self.semesters.keys()):
			if '.' in semester_year:
				continue
			ws = wb.create_sheet(semester_year, i)
			ws.cell(row=1, column=1, value='Semester')
			ws.cell(row=1, column=2, value='Year')
			ws.cell(row=1, column=3, value='Is term')
			ws.cell(row=1, column=4, value='Short Title')
			ws.cell(row=1, column=5, value='Number')
			ws.cell(row=1, column=6, value='Dept')
			ws.cell(row=1, column=7, value='Long title')
			ws.cell(row=1, column=8, value='Description')
			ws.cell(row=1, column=9, value='College short')
			ws.cell(row=1, column=10, value='College long')
			ws.cell(row=1, column=11, value='Course Offered')
			ws.cell(row=1, column=12, value='Course Headers')
			ws.cell(row=1, column=13, value='Course Note')
			ws.cell(row=1, column=14, value='When Taught')
			ws.cell(row=1, column=15, value='Prereqs')
			ws.cell(row=1, column=16, value='Recommended')
			ws.cell(row=1, column=17, value='Course level')
			ws.cell(row=1, column=18, value='Number of sections')
			ws.cell(row=1, column=19, value='Section number')
			ws.cell(row=1, column=20, value='Section type')
			ws.cell(row=1, column=21, value='Instructor')
			ws.cell(row=1, column=22, value='Credits')
			ws.cell(row=1, column=23, value='Term')
			ws.cell(row=1, column=24, value='Days')
			ws.cell(row=1, column=25, value='Starts')
			ws.cell(row=1, column=26, value='Ends')
			ws.cell(row=1, column=27, value='Location')
			ws.cell(row=1, column=28, value='Available')
			ws.cell(row=1, column=29, value='Seats')
			ws.cell(row=1, column=30, value='Waitlist')
			ws.cell(row=1, column=31, value='Building')

			cur_row = 2
			for course in self.semesters[semester_year].courses:
				for section in course.sections:
					ws.cell(row=cur_row, column=1, value=self.semesters[semester_year].semester)
					ws.cell(row=cur_row, column=2, value=self.semesters[semester_year].year)
					ws.cell(row=cur_row, column=3, value=self.semesters[semester_year].is_term)
					ws.cell(row=cur_row, column=4, value=course.short_title)
					ws.cell(row=cur_row, column=5, value=course.num)
					ws.cell(row=cur_row, column=6, value=course.dept)
					ws.cell(row=cur_row, column=7, value=course.long_title)
					ws.cell(row=cur_row, column=8, value=course.description)
					ws.cell(row=cur_row, column=9, value=course.college_short)
					ws.cell(row=cur_row, column=10, value=course.college_long)
					ws.cell(row=cur_row, column=11, value=course.offered)
					ws.cell(row=cur_row, column=12, value=course.headers)
					ws.cell(row=cur_row, column=13, value=course.note)
					ws.cell(row=cur_row, column=14, value=course.when_taught)
					ws.cell(row=cur_row, column=15, value=course.prerequisites)
					ws.cell(row=cur_row, column=16, value=course.recommended)
					ws.cell(row=cur_row, column=17, value=course.level)
					ws.cell(row=cur_row, column=18, value=course.num_sections)
					ws.cell(row=cur_row, column=19, value=section.section_num)
					ws.cell(row=cur_row, column=20, value=section.type)
					ws.cell(row=cur_row, column=21, value=section.instructor)
					ws.cell(row=cur_row, column=22, value=section.credits)
					ws.cell(row=cur_row, column=23, value=section.term)
					ws.cell(row=cur_row, column=24, value='\n'.join(section.days))
					ws.cell(row=cur_row, column=25, value='\n'.join(section.start))
					ws.cell(row=cur_row, column=26, value='\n'.join(section.end))
					ws.cell(row=cur_row, column=27, value=section.loction)
					ws.cell(row=cur_row, column=28, value=section.available)
					ws.cell(row=cur_row, column=29, value=section.seats)
					ws.cell(row=cur_row, column=30, value=section.waitlist)
					ws.cell(row=cur_row, column=31, value=section.building)
					cur_row += 1

		wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
		wb.save('semesters.xlsx')


	def semester(self, semester_year):
		if semester_year in self.semesters.keys():
			return self.semesters[semester_year]
		print(f'Requested semester {semester_year} not in semester manager')
		return None

	def update(self, semester_year):
		self.semesters[semester_year] = semester_getter.get(semester_year)
		if os.path.exists(os.path.join(self.pickles_path, semester_year)):
			os.rename(os.path.join(self.pickles_path, semester_year), os.path.join(self.pickles_path, semester_year + '.bak'))
		if self.semesters[semester_year] is not None:
			pickle.dump(self.semesters[semester_year], open(os.path.join(self.pickles_path, semester_year), 'wb'))

	def cached_semesters(self):
		return list(self.semesters.keys())

	def select_semester(self, semester_year):
		self.selected_semester = self.semesters[semester_year]

	def num_sections(self):
		return len(self.filtered_sections)

	def find_same_course(self, list_in, course):
		for i in range(len(list_in)):
			if course.short_title == list_in[i].short_title:
				if 'R' in course.short_title:
					print(f'Found an R course in {course.short_title}')
				return i

	def del_section(self, sections, section_num):
		for section in sections:
			if section.section_num == section_num:
				sections.remove(section)
				return None
		print('Did not find a section')

	def get_sections_by_course(self, courses, short_name):
		for course in courses:
			if course.short_title == short_name:
				return course.sections
		print('Did not find a course')

	def class_in_slot(self, day, section):
		for slot in section.schedule:
			if slot[0] != day:
				continue
			if slot[1] >= self.day_filter[day][0] and slot[2] <= self.day_filter[day][1]:
				return True
		return False

	def get_filtered_semester(self):
		semester_attributes = {'timestamp': self.selected_semester.timestamp,
		              'datestamp': self.selected_semester.datestamp,
		              'courses': [],
		              'semester year': self.selected_semester.semester_year}

		for course in self.selected_semester.courses:
			course_attributes = {'dept': course.dept,
			                     'num': course.num,
			                     'long title': course.long_title,
			                     'description': course.description,
			                     'college short': course.college_short,
			                     'college long': course.college_long,
			                     'courseCredits': course.hours,
			                     'courseOffered': course.offered,
			                     'courseHeaders': course.headers,
			                     'courseNote': course.note,
			                     'courseWhenTaught': course.when_taught,
			                     'coursePrereqs': course.prerequisites,
			                     'courseRec': course.recommended,
			                     'sections': []}
			if self.dept_filter not in course.dept:
				continue

			elif self.course_num_filter not in str(course.num):
				continue

			elif self.course_name_filter not in course.long_title:
				continue

			elif True in self.course_level_filter.values() and not self.course_level_filter[course.level]:
				continue

			for section in course.sections:
				make_section = True
				if self.instructor_filter not in section.instructor:
					make_section = False

				elif True in self.type_filter.values() and not self.type_filter[section.type]:
					make_section = False

				elif self.credits_filter[0] != 0 or self.credits_filter[1] != 0:
					if not (self.credits_filter[0] <= section.credits <= self.credits_filter[1]):
						make_section = False

				for day in self.day_filter.keys():
					if self.day_filter[day][2]:
						if not self.class_in_slot(day, section):
							make_section = False

				if make_section:
					section_attributes = {'section num': section.section_num,
					                      'type': section.type,
					                      'instructor': section.instructor,
					                      'credits': str(section.credits),
					                      'term': section.term,
					                      'days': section.days,
					                      'starts': section.start,
					                      'ends': section.end,
					                      'loction': section.loction,
					                      'available': section.available_frac,
					                      'waitlist': section.waitlist}
					course_attributes['sections'].append(classroom.section.Section(section_attributes))

			if len(course.sections) > 0:
				semester_attributes['courses'].append(classroom.course.Course(course_attributes))

		return classroom.semester.Semester(semester_attributes)

	def remove_semester(self, semester_year):
		del self.semesters[semester_year]
		if os.path.exists(semester_year):
			os.remove(semester_year)
